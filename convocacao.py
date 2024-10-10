# Descrever os passos manuais e transformar em código
# Ler a planilha e guardar informações sobre nome e telefone
import openpyxl
import webbrowser
import pyautogui
from urllib.parse import quote
from time import sleep

webbrowser.open('https://web.whatsapp.com/')
sleep(25)

workbook = openpyxl.load_workbook('servos.xlsx')
pagina = workbook['Planilha1']

workbook_erros = openpyxl.Workbook()
pagina_erros = workbook_erros.active
# Adiciona o cabeçalho, caso a planilha esteja sendo criada pela primeira vez
pagina_erros.append(['Nome Completo', 'Telefone', 'Célula', 'Email', 'Escala'])

for linha in pagina.iter_rows(min_row=2):
    nome_completo = linha[0].value
    nome = nome_completo.split()[0]
    telefone = linha[1].value
    celula = linha[2].value
    email = linha[3].value
    escala = linha[4].value
    mensagem = f'Olá, {nome}! Tudo bem?\nVi que você está escalado(a) no Salt pra servir nesse sábado, contamos com você.\nVocê pode confirmar pra mim se vai conseguir servir?\nÉ só responder nesse link aqui: https://forms.gle/mmXMidE7GbcpgM2n6'

    # Criar links personalizados do wpp e enviar mensagens para cada cliente com base nos dados da planilha
    if (escala == ''):
        try:
            link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
            webbrowser.open(link_mensagem)
            sleep(7)
            seta = pyautogui.locateCenterOnScreen('enter.png')
            sleep(4)
            pyautogui.click(seta[0], seta[1])
            sleep(4)
            pyautogui.hotkey('ctrl', 'w')
            sleep(4)
        except:
            print(f'Não foi possível enviar a mensagem para {nome_completo}')
            pagina_erros.append([nome_completo, telefone, celula, email, escala])

workbook_erros.save('erros.xlsx')