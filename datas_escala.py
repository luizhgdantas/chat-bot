# Descrever os passos manuais e transformar em código
# Ler a planilha e guardar informações sobre nome e telefone
import openpyxl
import webbrowser
import pyautogui
from urllib.parse import quote
from time import sleep

webbrowser.open('https://web.whatsapp.com/')
sleep(25)

workbook = openpyxl.load_workbook('servos.xlsx')
pagina = workbook['Planilha1']

workbook_erros = openpyxl.Workbook()
pagina_erros = workbook_erros.active
# Adiciona o cabeçalho, caso a planilha esteja sendo criada pela primeira vez
pagina_erros.append(['Nome Completo', 'Telefone', 'Célula', 'Email', 'Escala'])

for linha in pagina.iter_rows(min_row=2):
    nome_completo = linha[0].value
    nome = nome_completo.split()[0]
    telefone = linha[1].value
    celula = linha[2].value
    email = linha[3].value
    escala = linha[4].value

    mensagem1 = f'Oi, {nome}! Como você está?\nAcabamos de atualizar a escala no link da descrição do grupo do Salt, passando só pra te informar os dias que você estará escalado(a) pra que você possa se preparar e separar na sua agenda.\nVocê está escalado(a) no(s) dia(s): 12 de outubro'
    mensagem2 = f'Oi, {nome}! Como você está?\nAcabamos de atualizar a escala no link da descrição do grupo do Salt, passando só pra te informar os dias que você estará escalado(a) pra que você possa se preparar e separar na sua agenda.\nVocê está escalado(a) no(s) dia(s): 19 de outubro'
    mensagem3 = f'Oi, {nome}! Como você está?\nAcabamos de atualizar a escala no link da descrição do grupo do Salt, passando só pra te informar os dias que você estará escalado(a) pra que você possa se preparar e separar na sua agenda.\nVocê está escalado(a) no(s) dia(s): 26 de outubro'
    mensagem4 = f'Oi, {nome}! Como você está?\nAcabamos de atualizar a escala no link da descrição do grupo do Salt, passando só pra te informar os dias que você estará escalado(a) pra que você possa se preparar e separar na sua agenda.\nVocê está escalado(a) no(s) dia(s): 2 de novembro'

    # print(nome)
    # print(telefone)
    # print(celula)
    # https://web.whatsapp.com/send?phone=&text

    # Criar links personalizados do wpp e enviar mensagens para cada cliente com base nos dados da planilha
    try:
        if (escala == '1'):
            link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem1)}'
            webbrowser.open(link_mensagem)
            sleep(7)
            seta = pyautogui.locateCenterOnScreen('enter.png')
            sleep(4)
            pyautogui.click(seta[0], seta[1])
            sleep(4)
            pyautogui.hotkey('ctrl', 'w')
            sleep(4)
        elif (escala == '4'):
            link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem2)}'
            webbrowser.open(link_mensagem)
            sleep(7)
            seta = pyautogui.locateCenterOnScreen('enter.png')
            sleep(4)
            pyautogui.click(seta[0], seta[1])
            sleep(4)
            pyautogui.hotkey('ctrl', 'w')
            sleep(4)
        elif (escala == '3'):
            link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem3)}'
            webbrowser.open(link_mensagem)
            sleep(7)
            seta = pyautogui.locateCenterOnScreen('enter.png')
            sleep(4)
            pyautogui.click(seta[0], seta[1])
            sleep(4)
            pyautogui.hotkey('ctrl', 'w')
            sleep(4)
        elif (escala == '2'):
            link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem4)}'
            webbrowser.open(link_mensagem)
            sleep(7)
            seta = pyautogui.locateCenterOnScreen('enter.png')
            sleep(4)
            pyautogui.click(seta[0], seta[1])
            sleep(4)
            pyautogui.hotkey('ctrl', 'w')
            sleep(4)
    except:
        print(f'Não foi possível enviar a mensagem dos dias de escala para {nome_completo}')
        pagina_erros.append([nome_completo, telefone, celula, email, escala])

workbook_erros.save('erros_dias.xlsx')