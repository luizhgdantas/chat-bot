import customtkinter as ctk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from urllib.parse import quote
from time import sleep
import threading  # Para não travar a interface durante o envio

# Configuração do customtkinter
ctk.set_appearance_mode("dark")  # Modo escuro
ctk.set_default_color_theme("blue")  # Tema azul

# Variáveis globais
caminho_planilha = ""
escala_input = ""
ministerio_selecionado = ""

# Função para enviar mensagens usando Selenium
def iniciar_envio(planilha_path, escala_criterio, ministerio):
    try:
        # Inicializa o ChromeDriver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service)

        # Carregar planilha com informações de contato
        workbook = openpyxl.load_workbook(planilha_path)
        pagina = workbook['Planilha1']

        # Criar uma nova planilha para registrar erros
        workbook_erros = openpyxl.Workbook()
        pagina_erros = workbook_erros.active
        pagina_erros.append(['Nome Completo', 'Telefone', 'Célula', 'Email', 'Escala'])

        # Acessar o WhatsApp Web
        driver.get('https://web.whatsapp.com/')
        sleep(30)  # Tempo para escanear o QR code

        # Percorrer os contatos na planilha
        for i, linha in enumerate(pagina.iter_rows(min_row=2), start=1):
            nome_completo = linha[0].value
            nome = nome_completo.split()[0] if nome_completo else "Amigo(a)"
            telefone = linha[1].value
            celula = linha[2].value
            email = linha[3].value
            escala = linha[4].value
            mensagem = (f'Olá, {nome}! Tudo bem? Vi que você está escalado(a) para servir no "{ministerio}" ' 
                        f'neste sábado. Contamos com você. Confirme pelo link: https://forms.gle/mmXMidE7GbcpgM2n6')

            # Verifica se o critério da escala é atendido
            if str(escala) == escala_criterio:
                try:
                    # Criar link de mensagem no WhatsApp
                    link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
                    driver.get(link_mensagem)

                    # Aguardar botão de enviar
                    send_button = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[2]/button/span'))
                    )
                    send_button.click()
                    sleep(2)  # Pausa para garantir envio

                except Exception as e:
                    print(f"Erro ao enviar para {nome_completo}: {e}")
                    pagina_erros.append([nome_completo, telefone, celula, email, escala])

            # Atualizar status na interface
            lbl_status.configure(text=f"Enviando mensagem {i}/{len(list(pagina.iter_rows(min_row=2)))}...")
            janela_principal.update_idletasks()

        # Salvar erros, se houver
        workbook_erros.save('erros.xlsx')
        driver.quit()
        return "Envio concluído com sucesso!"
    except Exception as e:
        return f"Erro durante o envio: {e}"

# Função para selecionar a planilha
def selecionar_planilha():
    global caminho_planilha
    caminho_planilha = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if caminho_planilha:
        lbl_status.configure(text=f"Planilha selecionada: {caminho_planilha}")
    else:
        lbl_status.configure(text="Nenhuma planilha selecionada.")

# Função para executar o envio em uma thread separada
def enviar_thread():
    global escala_input
    escala_input = entrada_escala.get()
    if not caminho_planilha:
        messagebox.showerror("Erro", "Selecione uma planilha antes de iniciar.")
        return
    if not escala_input:
        messagebox.showerror("Erro", "Informe o número ou critério da escala.")
        return

    lbl_status.configure(text="Iniciando envio... Por favor, aguarde.")
    btn_iniciar.configure(state="disabled")  # Desabilitar botão durante o envio
    resultado = iniciar_envio(caminho_planilha, escala_input, ministerio_selecionado)
    messagebox.showinfo("Resultado", resultado)
    lbl_status.configure(text=resultado)
    btn_iniciar.configure(state="normal")  # Reabilitar botão

# Função para iniciar a janela principal
def abrir_janela_principal(ministerio):
    global ministerio_selecionado
    ministerio_selecionado = ministerio
    janela_selecao.destroy()

    global janela_principal
    janela_principal = ctk.CTk()
    janela_principal.title("Convocação Servos")
    janela_principal.geometry("600x400")

    # Componentes da interface principal
    lbl_titulo = ctk.CTkLabel(janela_principal, text=f"Convocação - {ministerio}", font=ctk.CTkFont(size=20, weight="bold"))
    lbl_titulo.pack(pady=10)

    btn_selecionar = ctk.CTkButton(janela_principal, text="Selecionar Planilha", command=selecionar_planilha)
    btn_selecionar.pack(pady=10)

    lbl_escala = ctk.CTkLabel(janela_principal, text="Informe o número/critério da escala:")
    lbl_escala.pack(pady=5)

    global entrada_escala
    entrada_escala = ctk.CTkEntry(janela_principal)
    entrada_escala.pack(pady=5)

    global lbl_status
    lbl_status = ctk.CTkLabel(janela_principal, text="Nenhuma planilha selecionada", wraplength=400)
    lbl_status.pack(pady=10)

    global btn_iniciar
    btn_iniciar = ctk.CTkButton(janela_principal, text="Iniciar Envio", command=lambda: threading.Thread(target=enviar_thread).start())
    btn_iniciar.pack(pady=20)

    janela_principal.mainloop()

# Janela inicial para selecionar ministério
janela_selecao = ctk.CTk()
janela_selecao.title("Seleção de Ministério")
janela_selecao.geometry("600x400")

lbl_escolha = ctk.CTkLabel(janela_selecao, text="Escolha o Ministério:", font=ctk.CTkFont(size=16, weight="bold"))
lbl_escolha.pack(pady=20)

ministerios = ["Salt operacional", "Ordem de culto", "Louvor", "Técnica", "Join"]
for ministerio in ministerios:
    btn_ministerio = ctk.CTkButton(janela_selecao, text=ministerio, command=lambda m=ministerio: abrir_janela_principal(m))
    btn_ministerio.pack(pady=10)

janela_selecao.mainloop()