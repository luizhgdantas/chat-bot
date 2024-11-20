from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from urllib.parse import quote
from time import sleep
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Inicializa o ChromeDriver usando o Service
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# Carregar planilha com informações de contato
workbook = openpyxl.load_workbook('servos.xlsx')
pagina = workbook['Planilha1']

# Criar uma nova planilha para registrar erros, se necessário
workbook_erros = openpyxl.Workbook()
pagina_erros = workbook_erros.active
pagina_erros.append(['Nome Completo', 'Telefone', 'Célula', 'Email', 'Escala'])

# Acessar o WhatsApp Web
driver.get('https://web.whatsapp.com/')
sleep(20)  # Dê um tempo para escanear o QR code manualmente

# Percorrer os contatos na planilha
for linha in pagina.iter_rows(min_row=2):
    nome_completo = linha[0].value
    nome = nome_completo.split()[0]
    telefone = linha[1].value
    celula = linha[2].value
    email = linha[3].value
    escala = linha[4].value
    mensagem = f'Olá, {nome}! Tudo bem?\nVi que você está escalado(a) no Salt pra servir nesse sábado, contamos com você.\nVocê pode confirmar pra gente se vai conseguir ou não servir?\nÉ só responder nesse link aqui: https://forms.gle/mmXMidE7GbcpgM2n6'

    if escala == '1':
        try:
            # Criar link de mensagem no WhatsApp
            link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
            driver.get(link_mensagem)

            # Aguardar até que o botão de enviar esteja clicável
            send_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[2]/button/span'))
            )
            send_button.click()  # Clica no botão de enviar
            sleep(2)  # Aguarda o envio

        except Exception as e:
            print(f'Erro ao enviar mensagem para {nome_completo}: {str(e)}')
            pagina_erros.append([nome_completo, telefone, celula, email, escala])

# Salvar planilha de erros
workbook_erros.save('erros.xlsx')

# Fechar o navegador
driver.quit()